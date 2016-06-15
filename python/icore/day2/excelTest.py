import openpyxl

wb = openpyxl.load_workbook('test1.xlsx')
ws = wb.get_sheet_by_name('mysheet1')
ws['A2'] ='hello'
print( ws['A1'].value )
print( ws['A3'].value )
wb.save('test1.xlsx')


# wb = openpyxl.Workbook()
# ws1 = wb.active
# ws1.title ="mysheet1"
# ws2 = wb.create_sheet("mysheet2")
# 
# ws1['A1'] ='korea'
# ws1['A2'] = 20
# ws1.append([10,20,30] )
# ws1.append([100,200,300] )
# 
# ws2['A1'] ='hello'
# ws2['A2'] = 200
# wb.save('test1.xlsx')
# print("excel write")


# import xlsxwriter
# 
# wb = xlsxwriter.Workbook('text.xlsx')
# ws1 = wb.add_worksheet("mysheet1")
# ws2 = wb.add_worksheet("mysheet2")
# ws1.write('A1', 10)
# ws1.write( 1, 1, 100)
# 
# ws2.write('A1', 'hello')
# ws2.write('A2', 100)
# ws2.write('A3', 200)
# ws2.write('A4','=sum(A2:A3)')
# 
# wb.close()
# print("excel write...")